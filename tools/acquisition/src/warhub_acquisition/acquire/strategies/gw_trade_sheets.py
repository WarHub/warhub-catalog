"""GW trade-site spreadsheet strategy: manufacturer-authoritative EANs from trade.games-workshop.com.

Registered as `STRATEGIES["gw-trade-sheets"]`. Enumerates the PUBLIC media library on Games
Workshop's retailer-network site, downloads the barcode-bearing workbooks, and emits one
Observation per product row. See `docs/research/2026-07-22-gw-trade-barcode-retrieval.md` for the
live probe this is built from (endpoint mechanics, full file inventory, measured yield, and the
terms assessment).

**Why this source exists at all.** Every prior probe concluded GW publishes no EANs on its own web
properties (`docs/research/2026-07-12-source-probe-manufacturers.md`) and that its trade order
forms were auth-walled (`2026-07-16-trade-order-sheets.md`). Both are wrong: the trade site serves
spreadsheets carrying retail EAN-13s from `/assets/YYYY/MM/<name>`, listed by an unauthenticated
REST route. `robots.txt` is fully open (`User-agent: *` / empty `Disallow:`).

## Enumeration

`GET {baseUrl}/wp-json/gw/v2/media?fe=1&type=118&order=desc&per_page=100&page=N&lang=en&country=C`
with an `X-WP-Nonce` header. Three non-obvious mechanics, each of which independently makes the
endpoint look empty or gated if you get it wrong:

1. **`country` must be the NUMERIC id** (220 = United Kingdom). Passing the country *name* returns
   `total_items: 0` with HTTP 200 -- indistinguishable from "no such data" and the single reason
   the previous probe reported a login wall.
2. **The nonce is public**, printed into `/resources/` HTML as `var gwAssetData = {"nonce":"..."}`.
   It is NOT a credential (requests succeed with no cookie and no session) but it IS required, and
   it rotates -- hence `_scrape_nonce`, which re-reads it at the start of every run rather than
   pinning a value in the descriptor.
3. **`type=118`** ("Printable Materials") is the documents bucket: ~468 items per country versus
   5,270 unfiltered, and every barcode-bearing spreadsheet found carries it. Enumerating unfiltered
   costs 10x the requests for the same yield.

Asset visibility is country-scoped and the slices genuinely differ (totals range 5,109-5,927), so
`scope.countries` fans out. Several barcode files exist ONLY in non-UK slices.

## Rate limiting -- the failure mode that silently truncates

The host 429s under load, and (verified 2026-07-22) **the media API degrades to a HTTP 200 with an
empty `assets` array rather than returning 429**. A paginator that reads `len(assets)` therefore
concludes "end of results" and reports success having collected a fraction of the data. During the
investigation this was initially misdiagnosed as a hard 1,800-item pagination cap; it is not, and
pages 19-53 return data normally at >=8s spacing.

`_fetch_page` therefore treats "empty assets before `total_items` is satisfied" as a RETRYABLE
THROTTLE, not as end-of-results, and gives up loudly (FetchError) rather than quietly. This is the
single most important behaviour in this module -- without it the source under-reports and its
contract still passes.

## Rows -> Observations

`key = f"{descriptor.id}:{product_code}"`, `sku` = GW's 11-digit product code (the join key the
catalog already stores as `productCode`), `ean` = the row's barcode, `name` = the trade
description. `manufacturer` is PINNED to `scope.manufacturer` ("Games Workshop") -- trade sheets
carry no vendor column, same situation as `mfr-gw-algolia`/`arc-gw-webstore`, resolved through
`Taxonomy.manufacturer_for_vendor`.

**Deletions rows set `archived=True`.** That is the existing, code-free lever for discontinued
products: `resolve/attributes.py` derives `status="discontinued"` for an entity with no
non-archived member. It also means a Deletions row can never flip a fresh entity to `current`.

Three data hazards, each measured and gated (research doc SS4.2):

- **`_GS1_PREFIXES` allowlist is mandatory, not defensive.** ~85 rows carry 12-digit GW-INTERNAL
  codes (11-digit product code + check digit, e.g. `608899990183` for code `60889999018`).
  `ean.normalize_ean` zero-pads 12-digit input into EAN-13 as UPC-A, so these pass
  `canonical_ean` cleanly and would be stored as retail barcodes. Only `5011921` (GW's GS1 prefix)
  and `977`/`978`/`979` (Bookland, for Black Library ISBN-13s) are accepted.
- **14-digit `Barcode (6-Pack)` values are GTIN-14 case codes**, not retail barcodes. Never read
  that column; `Barcode (Single)` is the unit EAN.
- **Future-dated rows are dropped** (`_release_date_is_future`). GW's Trade Terms define
  Confidential Information to expressly include "product release dates" and unreleased product
  info; excluding not-yet-released rows keeps that class of data out of the catalog entirely. This
  is a deliberate policy gate, not a data-quality one -- see the research doc SS6.

Prices: only RRP columns are read. Measured against 822 overlapping catalog products, the regional
`UKR` column has median(sheet/catalog priceGbp) = 1.000, i.e. it IS the retail price. The separate
`Trade Price`/`Cost` columns are wholesale (~64-65% of RRP) and are never read.

`full_sweep` is always False: this is a budgeted slice of workbooks, not a population census of
GW's range, and must never drive miss-streak/liveness decisions for products it does not list.
"""
from __future__ import annotations

import datetime as _dt
import io
import json
import re
from pathlib import Path
from typing import Iterator

from warhub_acquisition.acquire.client import FetchError, PoliteClient
from warhub_acquisition.acquire.runner import STRATEGIES, AcquireContext, StrategyResult
from warhub_acquisition.ean import canonical_ean
from warhub_acquisition.models.descriptor import SourceDescriptor
from warhub_acquisition.models.observation import Observation

# GW's own GS1 company prefix, plus the Bookland prefixes that carry Black Library ISBN-13s.
# Anything else in a Barcode column is not a retail barcode -- see the module docstring.
_GS1_PREFIXES: tuple[str, ...] = ("5011921", "977", "978", "979")

# GW product codes are exactly this many digits; see _code for why the width is load-bearing.
_CODE_DIGITS = 11

_NONCE_RE = re.compile(r"gwAssetData\s*=\s*\{\s*\"nonce\"\s*:\s*\"([0-9a-f]+)\"")

# The media API's documents bucket ("Printable Materials").
_DOC_TYPE = 118

_SPREADSHEET_SUFFIXES = (".xlsx", ".xlsm")


def _scrape_nonce(client: PoliteClient, base_url: str, resources_path: str) -> str:
    """Read the public REST nonce out of the /resources/ HTML.

    Not a credential: the endpoint serves the same data to an anonymous client with no cookie. It
    does rotate, so it is re-read every run rather than pinned in the descriptor.
    """
    html = client.get_text(f"{base_url.rstrip('/')}{resources_path}")
    match = _NONCE_RE.search(html)
    if match is None:
        raise FetchError(
            f"{base_url}{resources_path}: no gwAssetData nonce in page HTML -- the resources page "
            "changed shape, or the response was an edge block rather than the real page",
            status=None,
        )
    return match.group(1)


def _fetch_page(
    client: PoliteClient,
    base_url: str,
    nonce: str,
    country: int,
    page: int,
    *,
    expect_more: bool,
) -> dict:
    """One media-API page, distinguishing a real empty page from a silent throttle.

    `expect_more` is True while the caller still has items outstanding per `total_items`. In that
    state an empty `assets` array CANNOT be end-of-results, so it is treated as a throttle and
    raised as a rate-limited FetchError (the runner's degraded-run path) rather than being accepted
    as "no more data". See the module docstring -- this is the difference between a short run that
    reports failure and a short run that reports success.
    """
    payload, _headers = client.get_json_response(
        f"{base_url.rstrip('/')}/wp-json/gw/v2/media",
        params={
            "fe": 1,
            "type": _DOC_TYPE,
            "order": "desc",
            "per_page": 100,
            "page": page,
            "lang": "en",
            "country": country,
        },
        headers={"X-WP-Nonce": nonce},
    )
    if not isinstance(payload, dict):
        raise FetchError(f"media page {page} (country {country}): non-object payload", status=None)
    assets = payload.get("assets")
    if not isinstance(assets, list):
        raise FetchError(f"media page {page} (country {country}): no assets array", status=None)
    if not assets and expect_more:
        raise FetchError(
            f"media page {page} (country {country}): empty assets array while items remain "
            "outstanding -- GW's edge degrades to an empty 200 under load instead of 429; "
            "treating as a throttle rather than end-of-results",
            status=200,
            rate_limited=True,
        )
    return payload


def _enumerate_assets(
    client: PoliteClient, base_url: str, nonce: str, countries: list[int]
) -> dict[str, dict]:
    """Union of type-118 documents across every configured country slice, keyed by file_url."""
    found: dict[str, dict] = {}
    for country in countries:
        page = 1
        seen = 0
        total: int | None = None
        while True:
            payload = _fetch_page(
                client,
                base_url,
                nonce,
                country,
                page,
                expect_more=total is not None and seen < total,
            )
            if total is None:
                raw_total = payload.get("total_items")
                total = int(raw_total) if isinstance(raw_total, int) else 0
            assets = payload.get("assets") or []
            if not assets:
                break
            for asset in assets:
                url = asset.get("file_url")
                if isinstance(url, str) and url:
                    found.setdefault(url, asset)
            seen += len(assets)
            if seen >= total:
                break
            page += 1
    return found


def _select_workbooks(assets: dict[str, dict], patterns: list[str]) -> list[tuple[str, dict]]:
    """Assets whose file name matches a configured pattern and is a readable workbook.

    Patterns are matched case-insensitively against the file name. Legacy `.xls` (BIFF) is
    deliberately NOT selected: openpyxl cannot read it, the only such file on the site is a
    price-change list with no unique barcodes, and silently skipping it inside the parser would
    look like a parse failure rather than a deliberate exclusion.
    """
    compiled = [re.compile(p, re.IGNORECASE) for p in patterns]
    out: list[tuple[str, dict]] = []
    for url, asset in sorted(assets.items()):
        name = str(asset.get("file_name") or url.rsplit("/", 1)[-1])
        if not name.lower().endswith(_SPREADSHEET_SUFFIXES):
            continue
        if any(rx.search(name) for rx in compiled):
            out.append((url, asset))
    return out


# A row is the header iff it contains one of these exact cell values. Detecting the header by a
# known token rather than "first non-trivial row" is what lets the same parser handle the three
# different banner layouts on the site: the AU/NZ price files bury the header under a paragraph of
# RRP small print, and the China Order Form buries it under two rows of "Releases For Next Week" /
# order-total labels that would themselves pass a naive >=3-non-empty-cells test.
_HEADER_TOKENS: frozenset[str] = frozenset(
    {
        "Product Code",
        "New Product Code",
        "Barcode",
        "Barcode (Single)",
        "New Individual barcode",
        "New Barcode",
        "PRODUCT NAME",
        "Unit Code",
        "Individual Code",
    }
)


# Every column this module reads, and the ONLY columns the committed snapshot may carry. The
# snapshot is a NORMALIZED EXTRACT, never the raw workbook: a raw copy would drag GW's wholesale
# `Trade Price`/`Cost` columns into git alongside the retail data we actually publish. Keeping the
# whitelist next to the readers means a new `_first(row, ...)` spelling that is not listed here
# fails visibly on a snapshot re-parse rather than silently reading None.
_CONSUMED_COLUMNS: frozenset[str] = frozenset(
    {
        # identity
        "Product Code", "New Product Code", "Unit Code", "Individual Code", "New SKU",
        "Description", "Description (ENG)", "PRODUCT NAME", "Product Description", "Product Name",
        "Barcode (Single)", "New Individual barcode", "Barcode", "New Barcode",
        # attributes / classification
        "SS Code", "SSC", "New SS Code", "Short Code",
        "Category (ENG)", "Range", "Trade range",
        "SIZE", "UKR",
        # release-date policy gate
        "Release Date (Global)", "Release Date", "Release Date (China)",
        # re-coding lineage
        "Old Product Code", "Old Code", "Previous Product Code", "Old SKU", "Original SKU",
        "Old Barcode", "Old Individual barcode", "Previous Barcode",
        "Date", "Change Date", "Effective Date",
        # Not read by this module (yet) but whitelisted deliberately: GW's SS Code is the product's
        # identity ACROSS a re-code -- the register keeps it on almost every pair -- so `Old SSC
        # Code` beside `New SS Code` is what distinguishes a real predecessor from a stale regional
        # edition left in the Old columns. Adjudicating the `conflicting` bucket needs it, and it was absent from
        # every snapshot row because a whitelist only keeps what it is told to. A column the parser
        # does not read yet is exactly the kind a durable extract must not throw away.
        "Old SSC Code", "Old SS Code", "Old SSC",
    }
)


def _cell(value: object) -> object:
    """A cell value as JSON. Dates become ISO strings, which `_as_date` already accepts, so a
    snapshot round-trip parses identically to the live workbook it came from."""
    if isinstance(value, _dt.datetime):
        return value.date().isoformat()
    if isinstance(value, _dt.date):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _extract(row: dict) -> dict:
    return {
        name: _cell(value)
        for name, value in row.items()
        if name in _CONSUMED_COLUMNS and value not in (None, "")
    }


def _live_rows(client, workbooks, stats: dict[str, int], load_workbook) -> Iterator[tuple[str, str, dict]]:
    """`(workbook name, sheet title, row)` from the network, one workbook at a time."""
    for url, asset in workbooks:
        try:
            payload = client.get_response(url).content
            book = load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
        except (FetchError, Exception):  # noqa: BLE001 - a bad workbook must not fail the run
            stats["parse_errors"] += 1
            continue
        stats["workbooks"] += 1
        name = str(asset.get("file_name") or url.rsplit("/", 1)[-1])
        try:
            for sheet in book.worksheets:
                for row in _rows(sheet):
                    yield name, sheet.title, row
        finally:
            book.close()


def _snapshot_rows(path: Path, stats: dict[str, int]) -> Iterator[tuple[str, str, dict]]:
    """The same triples, replayed from the committed extract in their original order.

    Order matters: `_merge` resolves a product code seen in several workbooks by first-wins/richest
    -wins rules, so replaying out of order would not reproduce the live result.
    """
    if not path.exists():
        raise FileNotFoundError(
            f"no committed snapshot at {path}; run `warhub-data acquire --source mfr-gw-trade` "
            "once (with network) to write one"
        )
    books: set[str] = set()
    with path.open(encoding="utf-8") as handle:
        for line in handle:
            if not line.strip():
                continue
            payload = json.loads(line)
            books.add(payload["workbook"])
            yield payload["workbook"], payload["sheet"], payload["row"]
    stats["workbooks"] = len(books)


def _write_snapshot(path: Path, rows: list[tuple[str, str, dict]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for workbook, sheet, row in rows:
            handle.write(
                json.dumps(
                    {"workbook": workbook, "sheet": sheet, "row": row},
                    sort_keys=True, separators=(",", ":"), ensure_ascii=False,
                )
                + "\n"
            )


def _rows(sheet) -> Iterator[dict]:
    """Header-keyed rows, locating the header by a known column token.

    See `_HEADER_TOKENS`: the header is the first row containing one of them, so banner/label rows
    above it (which vary per sheet family) are skipped regardless of how many cells they fill.
    """
    header: list[str] | None = None
    for raw in sheet.iter_rows(values_only=True):
        cells = ["" if c is None else str(c).strip() for c in raw]
        if header is None:
            if any(c in _HEADER_TOKENS for c in cells):
                header = cells
            continue
        if not any(cells):
            continue
        yield dict(zip(header, raw))


def _volume_ml(size: object, name: str) -> int | None:
    """Unit volume in millilitres, from the SIZE column (`12ml`, `400ml`) or the product name
    (`... (12ML) ...`, `MEPHISTON RED 12ML ...`). None when there is no ml figure (most products,
    and brushes/tools/books which have no volume)."""
    for text in (str(size or ""), name or ""):
        m = re.search(r"(\d{1,4})\s*ml\b", text, re.IGNORECASE)
        if m:
            return int(m.group(1))
    return None


def _row_hints(row: dict, sheet_title: str, name: str) -> dict[str, object]:
    """The taxonomy a trade row carries, VERBATIM: its short code, its trade range, the sheet it
    lives on, and its unit volume.

    NO CATEGORY IS STAMPED HERE, since 2026-09-02. Until then this function's predecessors wrote
    `category: paint` for a `Paint -`/`Spray` range or a `Paint` sheet, which was a rule table
    hidden inside a harvester: a source's own claim outranks every table downstream
    (resolve/attributes.py), so the judgement could never be measured, argued with or corrected
    where the other sources' judgements live. It now lives in
    data/catalog/taxonomy/category-rules/mfr-gw-trade.yaml, keyed on exactly the two values this
    function records. Measured at the move: 938 rows carried the stamp, 324 of them from the range
    and 614 from the sheet, and the table decides every one the same way with basis `mapped`
    instead of `stated`.

    `sheets` IS A LIST because a product appears in several workbooks (the InsertDelete register,
    the Trade Direct Range, the WH Colour rebrand file) and `_merge` unions list hints across them
    -- the rebrand workbook's `Range` holds merchandising codes ("BS:A") and its tabs, `Paint` /
    `Brushes` / `Hobby`, are the only taxonomy it has, so that tab must survive a merge with a row
    from a workbook whose sheet is called `Sheet1`.
    """
    hints: dict[str, object] = {}
    ssc = _first(row, "SS Code", "SSC", "New SS Code", "Short Code")
    if ssc is not None:
        hints["sscCode"] = str(ssc).strip()
    category = _first(row, "Category (ENG)", "Range", "Trade range")
    if category is not None:
        hints["tradeCategory"] = str(category).strip()
    if sheet_title and sheet_title.strip():
        hints["sheets"] = [sheet_title.strip()]
    volume = _volume_ml(_first(row, "SIZE"), name)
    if volume is not None:
        hints["volumeMl"] = volume
    return hints


def _code(value: object, known: frozenset[str] = frozenset()) -> str:
    """A GW product code as text, restoring a leading zero Excel dropped.

    GW product codes are exactly 11 digits and several families start with one (`01…`–`04…` are
    EU-region codes, e.g. `01010299044`). A code column typed as NUMBER loses that zero on the way
    through the workbook, so the sheet hands us `1010299044`. That 10-digit string fails the
    taxonomy's `\\d{11}` pattern, which means the product resolves to a NAME-SLUG entity id and
    publishes with no `productCode` at all -- measured 2026-07-30: 60 of 7,531 trade observations,
    plus 2 retired codes in the re-coding register, where the same loss would fabricate a product
    code that never existed.

    `known` is every full-width code in the same harvest, and it is what keeps this from inventing
    codes. A short code is only a lost leading zero if it is not ALSO explainable as a full code
    that lost its LAST digit: GW's own sheets contain such typos (measured: `6024999960` on 3 rows
    of 2 workbooks, where the real code `60249999604` is asserted elsewhere with the same barcode).
    Padding that one produced `06024999960` -- a code that has never existed -- and moved a real
    product's barcode onto it, which is exactly what `report --ean-guard` flagged. When a short code
    looks truncated it is left alone, failing the taxonomy pattern as it always did.
    """
    text = re.sub(r"\s+", "", str(value))
    if not text.isdigit() or len(text) >= _CODE_DIGITS:
        return text
    if any(f"{text}{digit}" in known for digit in "0123456789"):
        return text  # a full code that lost its last digit, not one that lost a leading zero
    return text.zfill(_CODE_DIGITS)


def _first(row: dict, *names: str):
    for name in names:
        if name in row and row[name] not in (None, ""):
            return row[name]
    return None


def _row_code(row: dict) -> str | None:
    """The row's own product code as raw text, for the pre-pass that builds `known` (see _code)."""
    value = _first(row, "Product Code", "New Product Code", "Unit Code", "Individual Code", "New SKU")
    return None if value is None else re.sub(r"\s+", "", str(value))


def _predecessor(
    row: dict, run_date: str | None = None, known: frozenset[str] = frozenset()
) -> tuple[str | None, object, str | None, str | None]:
    """The (old product code, raw old barcode, change date, old SS Code) a `Code Changes` row
    renumbers FROM.

    Verified against the live register (InsertDelete18.05.2026.xlsx, sheet `Code Changes`), whose
    header is: New Product Code | Old Product Code | Description | New SS Code | Old SSC Code |
    New Barcode | Old Barcode | Trade Range | Date. Only that sheet family carries `Old …` columns,
    so a row without them yields (None, None, None) and every other sheet is untouched. Extra
    spellings are accepted for the same reason `_first` is used throughout this module: GW's column
    headings drift between workbook generations.

    `Date` is the day the renumbering took effect -- a genuine archival fact, and the only date in
    this data that is not a (confidential) forward-looking release date. A future-dated one is
    dropped on the same policy gate as unreleased products.
    """
    old_code = _first(row, "Old Product Code", "Old Code", "Previous Product Code", "Old SKU",
                      "Original SKU")
    old_barcode = _first(row, "Old Barcode", "Old Individual barcode", "Previous Barcode")
    code = _code(old_code, known) if old_code is not None else None

    changed_on = None
    changed = _as_date(_first(row, "Date", "Change Date", "Effective Date"))
    if changed is not None:
        today = None
        if run_date is not None:
            try:
                today = _dt.date.fromisoformat(run_date)
            except ValueError:
                today = None
        if today is None or changed <= today:
            changed_on = changed.isoformat()

    # GW's SS Code is the product's IDENTITY across a re-code -- the register keeps it on the
    # overwhelming majority of pairs. That makes `Old SSC Code` the discriminator between a real
    # predecessor and a stale regional edition parked in the Old columns. It is only sound BETWEEN
    # competing claimants, never as a veto on a lone edge: a real minority of already-declared
    # pairs legitimately renumber their SSC (Kor'sarro Khan 48-88 -> 55-24). See
    # classify/supersessions.py for how to re-derive that minority; the count once quoted here and
    # there was wrong from the start.
    old_ssc = _first(row, "Old SSC Code", "Old SS Code", "Old SSC")
    return (code or None), old_barcode, changed_on, (str(old_ssc).strip() if old_ssc else None)


def _clean_ean(raw) -> str | None:
    """Canonical EAN-13, gated on GW's GS1 prefixes.

    The prefix gate is what stops GW's 12-digit internal codes -- which `canonical_ean` happily
    zero-pads into a valid-looking EAN-13 -- from being stored as retail barcodes.
    """
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    digits = re.sub(r"[^0-9]", "", text)
    if len(digits) != 13:
        # 14-digit GTIN-14 case codes and 11/12-digit internal codes are not retail barcodes.
        return None
    ean = canonical_ean(digits)
    if ean is None or not ean.startswith(_GS1_PREFIXES):
        return None
    return ean


def _as_date(value) -> _dt.date | None:
    """A cell as a date. Accepts real date/datetime cells, ISO text, and GW's `DD/MM/YYYY`.

    The slash form is not cosmetic: a text-typed date column comes through as `01/03/2021`, which
    `fromisoformat` rejects, so the date was silently dropped. Measured 2026-07-30 over the
    committed register extract: 192 rows, losing the change date on every lineage pair they carry.
    Read as DAY-first (GW is a UK company and its ISO rows agree) -- and unambiguously so here:
    ZERO of those 192 rows have both leading fields <= 12, so none of them could be read as
    month-first even in principle. Release-date columns are unaffected: all 6,277 are ISO, so the
    confidentiality gate never depended on this.
    """
    if isinstance(value, _dt.datetime):
        return value.date()
    if isinstance(value, _dt.date):
        return value
    if isinstance(value, str):
        text = value.strip()
        try:
            return _dt.date.fromisoformat(text[:10])
        except ValueError:
            pass
        match = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{4})", text)
        if match:
            day, month, year = (int(part) for part in match.groups())
            try:
                return _dt.date(year, month, day)
            except ValueError:
                return None
    return None


def _release_date_is_future(row: dict, run_date: str) -> bool:
    """True when the row describes a product not yet released as of the run date.

    Policy gate, not a data-quality one: GW's Trade Terms name "product release dates" and
    unreleased product information as Confidential Information, so unreleased rows are excluded
    from the catalog entirely. Rows with no release-date column are never excluded by this.
    """
    released = _as_date(_first(row, "Release Date (Global)", "Release Date", "Release Date (China)"))
    if released is None:
        return False
    try:
        today = _dt.date.fromisoformat(run_date)
    except ValueError:
        return False
    return released > today


def _price(row: dict, *names: str) -> float | None:
    """An RRP column, or None. Never reads `Trade Price`/`Cost` -- those are wholesale."""
    value = _first(row, *names)
    if value is None:
        return None
    try:
        price = float(str(value).replace(",", "").strip())
    except ValueError:
        return None
    return price if price > 0 else None


def _merge(existing: Observation, fresh: Observation) -> Observation:
    """Fold a second row for the same product code into the observation already built for it.

    The same code legitimately appears in several workbooks -- the InsertDelete register lists it,
    and the current Trade Direct Range prices it -- and the naive last-wins dict assignment threw
    data away in both directions: `priceGbp` only exists on the Trade Direct Range rows, so
    whichever workbook happened to sort last (InsertDelete, having no price column) blanked it.

    Rules:
    - **Scalars: first non-null wins.** A later row must never overwrite a populated field with
      `None`. Where both are populated the earlier value is kept, so behaviour does not depend on
      filename sort order.
    - **`hints` merge**, so an SSC code from one sheet and a trade category from another both
      survive -- and a LIST hint (`sheets`) is unioned, so every tab a product was seen on is kept.

    `archived` is deliberately NOT decided here -- see `_is_discontinued`. It cannot be a pairwise
    fold: the correct answer depends on which KIND of sheet each sighting came from, which a
    two-observation merge cannot see.
    """
    merged = existing.model_copy(deep=True)
    for field in ("name", "sku", "ean", "priceGbp", "priceUsd", "priceEur", "url", "imageUrl"):
        if getattr(merged, field, None) is None:
            setattr(merged, field, getattr(fresh, field, None))
    for key, value in fresh.hints.items():
        if isinstance(value, list) and isinstance(merged.hints.get(key), list):
            merged.hints[key] = [*merged.hints[key], *(v for v in value if v not in merged.hints[key])]
        else:
            merged.hints.setdefault(key, value)
    return merged


def _sheet_role(sheet_title: str) -> str:
    """Classify a sheet as evidence of withdrawal, of current availability, or neither.

    This distinction is load-bearing and easy to get wrong. The InsertDelete workbook is a
    HISTORICAL REGISTER, not a snapshot of today's range:

    - `Deletions` -- the product left the trade range on the row's date. Evidence of withdrawal.
    - `Insertions` -- the product ENTERED the range on the row's date. This says nothing about
      whether it is still sold: 1,683 codes (measured, 2026-07-22) appear in BOTH Insertions and
      Deletions, i.e. they were added and later withdrawn, and are genuinely discontinued. Treating
      an Insertions row as evidence of currency wrongly revives every one of them.
    - `Code Changes` -- a renumbering record; likewise says nothing about current availability.
    - Everything else (Trade Direct Range `Sheet1`, the paint/brush sheets) is a CURRENT range
      listing: presence there means GW sells it today, which legitimately overrides a stale
      Deletions row for a re-introduced product.
    """
    title = sheet_title.strip().lower()
    if title == "deletions":
        return "withdrawn"
    if title in ("insertions", "code changes"):
        return "historical"
    return "current"


def _is_discontinued(roles: set[str]) -> bool:
    """A product is discontinued iff something withdrew it and nothing current still lists it."""
    return "withdrawn" in roles and "current" not in roles


def _mint_lineage_records(observations: dict[str, Observation], stats: dict[str, int]) -> None:
    """Give a RETIRED product code its own record when nothing else in the harvest observes it.

    GW's register asserts that a code existed and (usually) what its barcode was, but the trade
    sheets only LIST current codes -- so most retired codes have no row of their own and therefore
    no entity, no record, and no way for a scan of that old box to resolve. Measured 2026-07-30: 86
    such codes, 85 of them carrying a real barcode. Without this they are asserted and then lost,
    which is the exact failure the whole archival direction exists to stop.

    This is the first evidence in the pipeline DERIVED rather than observed, so it is deliberately
    narrow:

      * only when the code is claimed by exactly ONE surviving code -- a fan-out is either a
        regional split or a filler code, and neither is a 1:1 lineage fact;
      * only for a well-formed product code, so a stray barcode in a code column cannot mint a
        product;
      * never when the code is already observed -- a real row always wins, and because the key is
        the same, a later harvest that starts listing the code simply overwrites this record;
      * `archived=True`, so the existing lifecycle rule resolves it to `status: discontinued`
        rather than inventing a liveness claim.

    The name is the register row's own description, which describes the product as its SUCCESSOR is
    listed -- the two are the same product by the register's own assertion, but the record is marked
    `lineageDerived` so a consumer can tell a re-listed name from an observed one.
    """
    survivors: dict[str, set[str]] = {}
    seed: dict[str, tuple[Observation, dict]] = {}
    for observation in list(observations.values()):
        for entry in observation.hints.get("supersedes") or []:
            code = str(entry.get("productCode") or "")
            if not code:
                continue
            survivors.setdefault(code, set()).add(observation.key)
            seed.setdefault(code, (observation, entry))

    for code, (survivor, entry) in sorted(seed.items()):
        source_id = survivor.key.split(":", 1)[0]
        key = f"{source_id}:{code}"
        if key in observations:
            continue  # observed in its own right; the real row is the better record
        if len(survivors[code]) != 1:
            stats["lineage_records_ambiguous"] += 1
            continue
        if not (code.isdigit() and len(code) == _CODE_DIGITS):
            stats["lineage_records_malformed"] += 1
            continue
        hints: dict[str, object] = {"lineageDerived": True}
        if entry.get("changedOn"):
            hints["retiredOn"] = entry["changedOn"]
        observations[key] = Observation(
            key=key,
            manufacturer=survivor.manufacturer,
            name=survivor.name,
            sku=code,
            ean=entry.get("ean"),
            hints=hints,
            firstSeen=survivor.firstSeen,
            lastSeen=survivor.lastSeen,
            archived=True,
            extractor="gw-trade-sheets-lineage",
        )
        stats["lineage_records"] += 1
        if entry.get("ean"):
            stats["lineage_records_with_barcode"] += 1


def _attach_lineage(
    lineage: list[tuple[str, str, object, str | None, str | None]],
    observations: dict[str, Observation],
    stats: dict[str, int],
) -> None:
    """Fold collected `Code Changes` rows into `hints["supersedes"]` on the surviving observation.

    Two filters, both load-bearing:

    1. **Placeholder old-barcodes.** GW reuses a handful of filler values in the `Old Barcode`
       column across entirely unrelated products -- measured 2026-07-22: 14 such values, one of
       them (`5011921182312`) on 29 different rows. Asserting those would fabricate barcode links
       between unrelated products and trip the resolver's shared-EAN detection.

       The test is whether one old barcode is claimed by MORE THAN ONE distinct old product code
       -- NOT whether it occurs more than once. The InsertDelete register is cumulative, so every
       genuine pairing is restated in each generation of the workbook (measured live: 2,729 rows
       carrying a predecessor across 3 workbooks for 919 real edges). Counting raw occurrences
       therefore rejects EVERY barcode, which is exactly what happened before this was measured.
    2. **GS1 prefix gate** (`_clean_ean`), identical to the primary barcode column -- GW's 12-digit
       internal codes parse as valid UPC-A and must never be stored as barcodes.

    The hint is a LIST: a product re-coded more than once accumulates several predecessors, and a
    single-element list keeps that case from needing a shape change later. Sorted for determinism.
    """
    codes_per_barcode: dict[str, set[str]] = {}
    cleaned: list[tuple[str, str, str | None, str | None, str | None]] = []
    for key, old_code, raw_barcode, changed_on, _old_ssc in lineage:
        old_ean = _clean_ean(raw_barcode)
        cleaned.append((key, old_code, old_ean, changed_on, _old_ssc))
        if old_ean is not None:
            codes_per_barcode.setdefault(old_ean, set()).add(old_code)

    by_key: dict[str, dict[str, dict[str, str]]] = {}
    for key, old_code, old_ean, changed_on, old_ssc in cleaned:
        if old_ean is not None and len(codes_per_barcode[old_ean]) > 1:
            stats["lineage_placeholder_barcodes"] += 1
            old_ean = None
        entry: dict[str, str] = {"productCode": old_code}
        if old_ean is not None:
            entry["ean"] = old_ean
        if changed_on is not None:
            entry["changedOn"] = changed_on
        if old_ssc:
            entry["ssc"] = old_ssc
        # Dedup on the old code: the same renumbering is restated across workbook generations.
        # A later row carrying a usable barcode upgrades an earlier bare entry.
        existing = by_key.setdefault(key, {}).get(old_code)
        if existing is None or ("ean" not in existing and "ean" in entry):
            by_key[key][old_code] = entry

    for key, entries in by_key.items():
        observation = observations.get(key)
        if observation is None:
            # The surviving code produced no observation (no valid barcode of its own, or it was
            # dropped as unreleased). Nothing to hang the lineage on -- counted, not invented.
            stats["lineage_unmatched"] += len(entries)
            continue
        observation.hints["supersedes"] = [entries[code] for code in sorted(entries)]
        stats["lineage_links"] += len(entries)
        stats["lineage_with_barcode"] += sum(1 for e in entries.values() if "ean" in e)


def gw_trade_sheets_strategy(
    descriptor: SourceDescriptor,
    client: PoliteClient,
    cursor: dict,
    context: AcquireContext,
) -> StrategyResult:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dependency is declared in pyproject
        raise RuntimeError(
            f"{descriptor.id}: openpyxl is required to parse GW trade workbooks "
            "(`uv sync` in tools/acquisition)"
        ) from exc

    stats = {
        "workbooks": 0,
        "rows": 0,
        "emitted": 0,
        "skipped_no_ean": 0,
        "skipped_bad_prefix": 0,
        "skipped_unreleased": 0,
        "parse_errors": 0,
        # Re-coding lineage captured from `Code Changes` rows (see _attach_lineage).
        "lineage_links": 0,
        "lineage_with_barcode": 0,
        "lineage_placeholder_barcodes": 0,
        "lineage_unmatched": 0,
        # Archival records minted for retired codes nothing observes (see _mint_lineage_records).
        "lineage_records": 0,
        "lineage_records_with_barcode": 0,
        "lineage_records_ambiguous": 0,
        "lineage_records_malformed": 0,
    }

    base_url = descriptor.baseUrl or "https://trade.games-workshop.com"
    scope = descriptor.scope
    # scope.manufacturer is the vendor NAME ("Games Workshop"); observations must carry the
    # taxonomy SLUG ("games-workshop"). Resolving through manufacturer_for_vendor is what makes
    # these rows join the existing GW entities -- emitting the raw name instead silently mints a
    # parallel 10th manufacturer and duplicates the whole GW catalog (measured: +7,999 products
    # and +7,157 conflicts on a resolve). Same mechanism as algolia.py/woo.py's pinned sources.
    manufacturer_name = str(scope.get("manufacturer") or "Games Workshop")
    manufacturer = context.taxonomy.manufacturer_for_vendor(manufacturer_name)
    if manufacturer is None:
        raise ValueError(
            f"{descriptor.id}: scope.manufacturer {manufacturer_name!r} is not a known vendor name "
            "in data/catalog/taxonomy/manufacturers.yaml (it must be a name/vendorName, not a slug)"
        )
    resources_path = str(scope.get("resourcesPath") or "/resources/")
    countries = [int(c) for c in (scope.get("countries") or [220])]  # type: ignore[union-attr]
    patterns = [str(p) for p in (scope.get("filePatterns") or [])]  # type: ignore[union-attr]
    if not patterns:
        raise ValueError(f"{descriptor.id}: scope.filePatterns is required")

    # Durability (phase 6): the workbooks are a LIVE third-party surface -- they rotate, get
    # re-uploaded under new names, and the `Code Changes` register is cumulative only for as long
    # as GW chooses to keep restating old rows. Every live run therefore writes a normalized
    # extract of exactly the columns this module reads, and `--from-snapshot` re-parses that
    # extract with no network at all. Re-deriving lineage must not depend on a 35-minute scrape of
    # someone else's site staying available.
    snapshot_path = (context.snapshot_dir / "sheets.jsonl") if context.snapshot_dir else None
    if context.from_snapshot:
        if snapshot_path is None:
            raise ValueError(f"{descriptor.id}: --from-snapshot needs a data root to read from")
        source_rows = _snapshot_rows(snapshot_path, stats)
        workbooks = []
    else:
        nonce = _scrape_nonce(client, base_url, resources_path)
        assets = _enumerate_assets(client, base_url, nonce, countries)
        workbooks = _select_workbooks(assets, patterns)
        if context.budget is not None:
            workbooks = workbooks[: context.budget]
        source_rows = _live_rows(client, workbooks, stats, openpyxl.load_workbook)
    # PRE-PASS. The release-date policy gate runs here rather than in the main loop so the snapshot
    # and the `known` code set are built from exactly the same rows on a live run and on a replay
    # -- otherwise a future-dated row would contribute a code live but not offline, and the two
    # would stop agreeing. `known` is what stops _code inventing a product code (see its docstring).
    run_date = context.run_date
    captured: list[tuple[str, str, dict]] = []
    known_codes: set[str] = set()
    for workbook_name, sheet_title, row in source_rows:
        stats["rows"] += 1
        if _release_date_is_future(row, run_date):
            # Dropped BEFORE capture: GW's Trade Terms make unreleased product information
            # Confidential, so a future release date may no more be committed to git than
            # published. The next live run picks the product up once it ships.
            stats["skipped_unreleased"] += 1
            continue
        captured.append((workbook_name, sheet_title, _extract(row)))
        raw_code = _row_code(row)
        if raw_code is not None and raw_code.isdigit() and len(raw_code) == _CODE_DIGITS:
            known_codes.add(raw_code)
    known = frozenset(known_codes)

    observations: dict[str, Observation] = {}
    # Per-product provenance across every sheet of every workbook, resolved into `archived` once
    # the whole harvest is in -- see _sheet_role / _is_discontinued.
    roles: dict[str, set[str]] = {}
    # (surviving observation key, old product code, raw old barcode) from `Code Changes` rows.
    # Collected here and resolved AFTER every workbook is parsed, because the placeholder filter
    # below can only be applied once all occurrences of an old barcode have been counted -- the
    # same deferred-decision shape as `roles`/_is_discontinued.
    lineage: list[tuple[str, str, object]] = []

    for _workbook_name, sheet_title, row in captured:
        role = _sheet_role(sheet_title)
        code = _first(row, "Product Code", "New Product Code", "Unit Code",
                      "Individual Code", "New SKU")
        name = _first(row, "Description", "Description (ENG)", "PRODUCT NAME",
                      "Product Description", "Product Name")
        if code is None or name is None:
            continue

        raw_barcode = _first(row, "Barcode (Single)", "New Individual barcode",
                             "Barcode", "New Barcode")
        ean = _clean_ean(raw_barcode)
        if ean is None:
            if raw_barcode is None:
                stats["skipped_no_ean"] += 1
            else:
                stats["skipped_bad_prefix"] += 1
            continue

        sku = _code(code, known)
        key = f"{descriptor.id}:{sku}"
        hints = _row_hints(row, sheet_title, str(name))

        # Re-coding lineage: this row says "old code -> this code". Recorded against
        # the SURVIVING key and resolved after the whole harvest (see below). A row
        # whose old code equals its new code is a no-op, not a supersession.
        old_code, old_barcode, changed_on, old_ssc = _predecessor(row, run_date, known)
        if old_code is not None and old_code != sku:
            lineage.append((key, old_code, old_barcode, changed_on, old_ssc))

        fresh = Observation(
            key=key,
            manufacturer=manufacturer,
            name=str(name).strip(),
            sku=sku,
            ean=ean,
            priceGbp=_price(row, "UKR"),
            hints=hints,
            firstSeen=run_date,
            lastSeen=run_date,
            # Provisional; finalised from `roles` once every workbook is parsed.
            archived=False,
            extractor="gw-trade-sheets",
        )
        existing = observations.get(key)
        observations[key] = fresh if existing is None else _merge(existing, fresh)
        roles.setdefault(key, set()).add(role)
        stats["emitted"] += 1

    # Refresh the committed extract only on a live run -- a `--from-snapshot` parse must never
    # rewrite the file it just read (a budgeted or partly-failed replay would truncate it).
    if snapshot_path is not None and not context.from_snapshot:
        _write_snapshot(snapshot_path, captured)
        stats["snapshot_rows"] = len(captured)

    for key, observation in observations.items():
        observation.archived = _is_discontinued(roles.get(key, set()))
    stats["discontinued"] = sum(1 for o in observations.values() if o.archived)

    _attach_lineage(lineage, observations, stats)
    # After the lineage hints exist, and after `archived` has been resolved from `roles` -- a
    # minted record sets its own `archived=True` and must not be overwritten by that pass.
    _mint_lineage_records(observations, stats)

    return StrategyResult(
        observations=list(observations.values()),
        # Never a population census: a budgeted slice of workbooks must not drive liveness.
        full_sweep=False,
        stats=stats,
        cursor=cursor,
    )


STRATEGIES["gw-trade-sheets"] = gw_trade_sheets_strategy
